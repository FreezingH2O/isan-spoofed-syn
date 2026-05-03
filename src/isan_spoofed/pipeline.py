from __future__ import annotations

import importlib
import sys
from pathlib import Path
from typing import Optional

import click
import yaml
from loguru import logger
from rich.console import Console
from rich.panel import Panel
from rich.table import Table
from rich.text import Text

console = Console()

_PHASES = {
    0: ("Dataset Analysis",           "isan_spoofed.phase0_analyze",  "run"),
    1: ("Phonetic Normalization",     "isan_spoofed.phase1_normalize", "run"),
    2: ("TTS Source Generation",      "isan_spoofed.phase2_tts",       "run"),
    3: ("Target Speaker Curation",    "isan_spoofed.phase3_curate",    "run"),
    4: ("RVC Model Training",         "isan_spoofed.phase4_train",     "run"),
    5: ("Voice Conversion Inference", "isan_spoofed.phase5_infer",     "run"),
    6: ("Audio Enhancement",          "isan_spoofed.phase6_enhance",   "run"),
}

# MODE SHORTCUTS
# tts  — text only, no voice conversion, no enhancement (F5-TTS output is already clean)
# rvc  — curate + train + infer + enhance (vocoder buzz is significant without TTS warmup)
# full — everything 0–6 (TTS→RVC stacks two synthesis stages; enhancement is most needed here)
_MODE_PHASES: dict[str, list[int]] = {
    "tts":  [1, 2],
    "rvc":  [3, 4, 5, 6],
    "full": [0, 1, 2, 3, 4, 5, 6],
}


def _load_config(config_path: Path) -> dict:
    with open(config_path) as f:
        return yaml.safe_load(f)


def _configure_logging(log_dir: Path, verbose: bool) -> None:
    log_dir.mkdir(parents=True, exist_ok=True)
    logger.remove()
    level = "DEBUG" if verbose else "INFO"
    logger.add(
        sys.stderr, level=level, colorize=True,
        format="<green>{time:HH:mm:ss}</green> | <level>{level: <8}</level> | {message}",
    )
    logger.add(log_dir / "pipeline.log", level="DEBUG", rotation="10 MB", retention="30 days")


def _print_phase_summary(phases: list[int]) -> None:
    table = Table(title="Execution Plan", show_header=True, header_style="bold cyan")
    table.add_column("Phase", style="dim", width=6)
    table.add_column("Name")
    for p in phases:
        table.add_row(str(p), _PHASES[p][0])
    console.print(table)


def _artifact_counts(cfg: dict) -> dict[str, int]:
    paths = cfg["paths"]
    checks = [
        ("Raw dataset wavs",     Path(paths["data"]["raw_dataset"]),            "*.wav"),
        ("Speakers curated",     Path(paths["data"]["speakers"]),               "*/"),
        ("Normalized texts",     Path(paths["data"]["normalized_text"]),        "*.txt"),
        ("TTS wav files",        Path(paths["data"]["tts_output"]),             "*.wav"),
        ("RVC models (.pth)",    Path(paths["models"]["rvc"]),                  "*.pth"),
        ("RVC indexes (.index)", Path(paths["models"]["rvc"]),                  "*.index"),
        ("Spoofed outputs",      Path(paths["output"]["spoofed"]),              "*.wav"),
        ("Enhanced outputs",     Path(paths["output"]["enhanced"]),             "*.wav"),
        ("Augmented outputs",    Path(paths["data"]["augmented"]),              "*.wav"),
        ("Speaker analysis",     Path(paths["analysis"]["speaker_xlsx"]).parent,"speaker_analysis.xlsx"),
    ]
    result = {}
    for label, base, pattern in checks:
        if pattern.endswith("/"):
            result[label] = sum(1 for p in base.iterdir() if p.is_dir()) if base.exists() else 0
        else:
            result[label] = len(list(base.glob(pattern))) if base.exists() else 0
    return result


def _execute_phases(
    cfg: dict,
    phase_list: list[int],
    speakers: Optional[str],
    extra: Optional[dict] = None,
) -> None:
    _print_phase_summary(phase_list)
    extra = extra or {}
    for phase_num in phase_list:
        name, module_path, fn_name = _PHASES[phase_num]
        logger.info(f"Starting Phase {phase_num}: {name}")
        try:
            module = importlib.import_module(module_path)
            fn = getattr(module, fn_name)
            # PHASE-SPECIFIC EXTRA PARAMS
            if phase_num == 1:
                fn(cfg, speaker_id=speakers,
                   text=extra.get("text"),
                   input_path=extra.get("input_path"))
            elif phase_num == 5:
                fn(cfg, speaker_id=speakers,
                   tts_stems=extra.get("tts_stems"),
                   source_dir=extra.get("source_dir"))
            else:
                fn(cfg, speaker_id=speakers)
            logger.success(f"Phase {phase_num} complete: {name}")
        except Exception as exc:
            logger.error(f"Phase {phase_num} failed: {exc}")
            raise click.ClickException(str(exc)) from exc


# INTERACTIVE MENU LOOP — SHOWN WHEN NO SUBCOMMAND IS GIVEN
def _run_menu(cfg: dict) -> None:
    while True:
        console.print()
        console.print(Panel(
            Text("ISAN SPOOFED — PIPELINE CONTROLLER", justify="center", style="bold white"),
            style="bold cyan",
            expand=False,
        ))
        console.print()

        # PHASE-BY-PHASE SECTION
        console.print("  [bold cyan]── Individual Phases ──[/bold cyan]")
        for num, (name, _, _) in _PHASES.items():
            console.print(f"    [bold yellow]{num}[/bold yellow]  {name}")

        console.print()
        console.print("  [bold cyan]── Pipeline Modes ──[/bold cyan]")
        console.print("    [bold yellow]t[/bold yellow]  TTS only         (phases 1–2: normalize → TTS audio, no enhance)")
        console.print("    [bold yellow]r[/bold yellow]  Pure RVC         (phases 3–4–5–6: curate → train → infer → enhance)")
        console.print("    [bold yellow]f[/bold yellow]  TTS → RVC        (phases 0–6: full pipeline with enhancement)")
        console.print()
        console.print("  [bold cyan]── Other ──[/bold cyan]")
        console.print("    [bold yellow]a[/bold yellow]  Augment output")
        console.print("    [bold yellow]s[/bold yellow]  Status")
        console.print("    [bold yellow]q[/bold yellow]  Quit")
        console.print()

        choice = click.prompt("  Choice", default="", show_default=False).strip().lower()

        # SINGLE PHASE
        if choice.isdigit() and int(choice) in _PHASES:
            phase_num = int(choice)
            kwargs = _prompt_phase_inputs(cfg, [phase_num])
            _execute_phases(cfg, [phase_num], kwargs.get("speakers"), extra=kwargs)

        # MODE SHORTCUTS
        elif choice == "t":
            kwargs = _prompt_phase_inputs(cfg, _MODE_PHASES["tts"])
            _execute_phases(cfg, _MODE_PHASES["tts"], kwargs.get("speakers"), extra=kwargs)
        elif choice == "r":
            kwargs = _prompt_phase_inputs(cfg, _MODE_PHASES["rvc"])
            _execute_phases(cfg, _MODE_PHASES["rvc"], kwargs.get("speakers"), extra=kwargs)
        elif choice == "f":
            kwargs = _prompt_phase_inputs(cfg, _MODE_PHASES["full"])
            _execute_phases(cfg, _MODE_PHASES["full"], kwargs.get("speakers"), extra=kwargs)

        # AUGMENT
        elif choice == "a":
            from isan_spoofed import augment as aug_module
            logger.info("Starting data augmentation")
            aug_module.run(cfg)
            logger.success("Augmentation complete")

        # STATUS
        elif choice == "s":
            counts = _artifact_counts(cfg)
            table = Table(title="Artifact Status", show_header=True, header_style="bold cyan")
            table.add_column("Artifact")
            table.add_column("Count", justify="right")
            for label, count in counts.items():
                table.add_row(label, str(count))
            console.print(table)
            click.pause(info="\nPress any key to continue...")

        elif choice in ("q", "quit", "exit"):
            console.print("[dim]Bye.[/dim]")
            break

        elif choice == "":
            continue

        else:
            console.print(f"[red]  Unknown choice: {choice!r}[/red]")


def _speakers_from_xlsx(cfg: dict) -> list[dict] | None:
    # READ SPEAKER LIST FROM PHASE 0 EXCEL — RETURNS None IF FILE NOT FOUND
    xlsx = Path(cfg["paths"]["analysis"]["speaker_xlsx"])
    if not xlsx.exists():
        return None
    try:
        from openpyxl import load_workbook
        wb = load_workbook(str(xlsx), read_only=True, data_only=True)
        ws = wb.active
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append(dict(zip(headers, row)))
        wb.close()
        return rows if rows else None
    except Exception:
        return None


def _speakers_from_raw_dataset(cfg: dict) -> list[dict] | None:
    # SCAN data/raw_dataset/ FOR SPEAKER IDs — RETURNS None IF EMPTY
    from isan_spoofed.utils.dataset import parse_speaker_id
    raw_dir = Path(cfg["paths"]["data"]["raw_dataset"])
    if not raw_dir.exists():
        return None
    counts: dict[str, int] = {}
    for f in raw_dir.rglob("*.wav"):
        sid = parse_speaker_id(f.name)
        if sid:
            counts[sid] = counts.get(sid, 0) + 1
    if not counts:
        return None
    return [{"speaker_id": sid, "gender": sid[0], "clip_count": n}
            for sid, n in sorted(counts.items())]


_TTS_LIST_LIMIT = 30  # SHOW INDIVIDUAL FILES BELOW THIS COUNT, FOLDERS ABOVE


def _tts_files_for_phase5(cfg: dict) -> dict | None:
    # SCAN tts_output/ AND RETURN {"mode": "files"|"folders", "items": [...]}
    tts_dir = Path(cfg["paths"]["data"]["tts_output"])
    if not tts_dir.exists():
        return None

    # COLLECT ALL WAVS INCLUDING SUBDIRECTORIES
    all_wavs = sorted(tts_dir.rglob("*.wav"))
    if not all_wavs:
        return None

    if len(all_wavs) <= _TTS_LIST_LIMIT:
        # SMALL COUNT — SHOW INDIVIDUAL FILES
        return {
            "mode": "files",
            "root": tts_dir,
            "items": [{"stem": w.stem, "filename": w.name, "_path": w} for w in all_wavs],
        }

    # LARGE COUNT — GROUP BY PARENT FOLDER
    groups: dict[Path, list[Path]] = {}
    for w in all_wavs:
        groups.setdefault(w.parent, []).append(w)

    items = []
    for folder, wavs in sorted(groups.items()):
        rel = folder.relative_to(tts_dir) if folder != tts_dir else Path(".")
        items.append({
            "folder": str(rel),
            "wav_count": len(wavs),
            "_path": folder,
            "_wavs": wavs,
        })

    return {"mode": "folders", "root": tts_dir, "items": items, "total": len(all_wavs)}


def _tts_picker(tts_data: dict) -> tuple[Optional[str], Optional[str]]:
    # RETURNS (tts_stems, source_dir) — EXACTLY ONE WILL BE SET, OR BOTH NONE (= ALL)
    mode = tts_data["mode"]
    items = tts_data["items"]
    root = tts_data["root"]

    if mode == "files":
        # SMALL COUNT: NUMBERED FILE LIST
        stems = _picker(items, id_col="stem",
                        title="TTS Source Files — Phase 5 (Inference)",
                        prompt="TTS file selection")
        return stems, None

    # FOLDER MODE: SHOW FOLDER TABLE
    total = tts_data["total"]
    table = Table(
        title=f"TTS Source Folders — {total} wav files total",
        show_header=True, header_style="bold cyan",
    )
    table.add_column("#", style="dim", width=4)
    table.add_column("folder", style="bold")
    table.add_column("wav_count", justify="right")

    for i, row in enumerate(items, start=1):
        table.add_row(str(i), row["folder"], str(row["wav_count"]))

    console.print()
    console.print(table)
    console.print()
    console.print("  Select source:")
    console.print("    [dim]• Numbers:  1,3     — specific folders[/dim]")
    console.print("    [dim]• Range:    1-5     — folder range[/dim]")
    console.print("    [dim]• Path:     ./my/folder  — any folder path[/dim]")
    console.print("    [dim]• All:      a (default)[/dim]")
    console.print()

    raw = click.prompt("  TTS source selection", default="a").strip()

    if not raw or raw.lower() == "a":
        return None, None  # ALL FILES

    # CHECK IF IT LOOKS LIKE A PATH
    if raw.startswith(".") or raw.startswith("/") or (len(raw) > 1 and raw[1] == ":"):
        p = Path(raw)
        if not p.exists():
            console.print(f"[red]  Path not found: {p}[/red]")
            return None, None
        return None, str(p)

    # PARSE AS ROW NUMBERS / RANGES
    selected_wavs: list[Path] = []
    for part in raw.split(","):
        part = part.strip()
        if "-" in part and part.replace("-", "").isdigit():
            start, _, end = part.partition("-")
            for i in range(int(start), int(end) + 1):
                if 1 <= i <= len(items):
                    selected_wavs.extend(items[i - 1]["_wavs"])
        elif part.isdigit():
            i = int(part)
            if 1 <= i <= len(items):
                selected_wavs.extend(items[i - 1]["_wavs"])

    if not selected_wavs:
        return None, None

    # IF SELECTION IS A SINGLE FOLDER, PASS AS source_dir FOR CLEAN LOGGING
    folders = {w.parent for w in selected_wavs}
    if len(folders) == 1:
        return None, str(folders.pop())

    # MULTI-FOLDER: PASS AS COMMA-SEPARATED STEMS
    stems = ",".join(w.stem for w in sorted(set(selected_wavs)))
    return stems, None


def _models_for_phase5(cfg: dict) -> list[dict] | None:
    # LIST TRAINED RVC MODELS IN models/rvc/
    model_dir = Path(cfg["paths"]["models"]["rvc"])
    if not model_dir.exists():
        return None
    pth_files = sorted(model_dir.glob("*.pth"))
    if not pth_files:
        return None
    rows = []
    for p in pth_files:
        has_index = (model_dir / f"{p.stem}.index").exists()
        rows.append({"speaker_id": p.stem, "index": "yes" if has_index else "[red]missing[/red]"})
    return rows


def _speakers_from_manifest(cfg: dict) -> list[dict] | None:
    # READ CURATED SPEAKERS FROM PHASE 3 MANIFEST
    import json
    manifest_path = Path(cfg["paths"]["data"]["speakers"]) / "manifest.json"
    if not manifest_path.exists():
        return None
    data: dict[str, str] = json.loads(manifest_path.read_text())
    if not data:
        return None
    return [{"speaker_id": sid, "wav": Path(p).name} for sid, p in sorted(data.items())]


def _parse_selection(raw: str, rows: list[dict], id_col: str = "speaker_id") -> Optional[str]:
    # PARSE SELECTION: "a"=all, "1,3" or "1-3"=row numbers, or direct IDs
    raw = raw.strip()
    if not raw or raw.lower() == "a":
        return None  # all items

    ids: list[str] = []
    for part in raw.split(","):
        part = part.strip()
        if "-" in part and part.replace("-", "").isdigit():
            start, _, end = part.partition("-")
            for i in range(int(start), int(end) + 1):
                if 1 <= i <= len(rows):
                    ids.append(str(rows[i - 1][id_col]))
        elif part.isdigit():
            i = int(part)
            if 1 <= i <= len(rows):
                ids.append(str(rows[i - 1][id_col]))
        elif part:
            ids.append(part)  # DIRECT ID TYPED BY USER

    return ",".join(dict.fromkeys(ids)) or None  # DEDUPLICATE, PRESERVE ORDER


def _picker(rows: list[dict], id_col: str, title: str, prompt: str = "Selection") -> Optional[str]:
    # DISPLAY NUMBERED TABLE AND RETURN SELECTED VALUES FOR id_col (None = all)
    table = Table(title=title, show_header=True, header_style="bold cyan")
    table.add_column("#", style="dim", width=4)

    other_cols = [k for k in rows[0].keys() if k != id_col]
    table.add_column(id_col, style="bold")
    for col in other_cols:
        table.add_column(col)

    for i, row in enumerate(rows, start=1):
        table.add_row(str(i), str(row[id_col]), *[str(row.get(c, "")) for c in other_cols])

    console.print()
    console.print(table)
    console.print()
    console.print(f"  Select {id_col.replace('_', ' ')}s:")
    console.print("    [dim]• Numbers:  1,3,5[/dim]")
    console.print("    [dim]• Range:    1-10[/dim]")
    console.print("    [dim]• Direct:   f_091,m_023[/dim]")
    console.print("    [dim]• All:      a (default)[/dim]")
    console.print()

    raw = click.prompt(f"  {prompt}", default="a").strip()
    return _parse_selection(raw, rows, id_col=id_col)


def _speaker_picker(rows: list[dict], title: str) -> Optional[str]:
    return _picker(rows, id_col="speaker_id", title=title)


def _prompt_phase_inputs(cfg: dict, phase_list: list[int]) -> dict:
    result: dict = {}

    # PHASE 1: ASK HOW TO SUPPLY TEXT INPUT
    if 1 in phase_list:
        input_dir = Path(cfg["paths"]["data"]["normalized_text"]) / "input"
        has_files = input_dir.exists() and any(input_dir.glob("*.txt"))

        console.print()
        console.print("  [bold cyan]── Phase 1 input source ──[/bold cyan]")
        console.print("    [bold yellow]1[/bold yellow]  Type / paste text now")
        console.print("    [bold yellow]2[/bold yellow]  Specify a file or folder path")
        if has_files:
            console.print(f"    [bold yellow]3[/bold yellow]  Use files already in [dim]{input_dir}[/dim]  (default)")
        console.print()

        default = "3" if has_files else "1"
        src = click.prompt("  Input source", default=default).strip()

        if src == "1":
            from isan_spoofed.phase1_normalize import _read_multiline_stdin
            result["text"] = _read_multiline_stdin()
        elif src == "2":
            result["input_path"] = click.prompt("  File or folder path").strip()

    # PHASE 3: SHOW AVAILABLE SPEAKERS TO CURATE
    if 3 in phase_list:
        rows = _speakers_from_xlsx(cfg) or _speakers_from_raw_dataset(cfg)
        if rows:
            result["speakers"] = _speaker_picker(rows, "Available Speakers — Phase 3 (Curation)")
        else:
            console.print()
            console.print("  [dim]No speaker list found. Run Phase 0 first, or drop WAVs in data/raw_dataset/.[/dim]")
            console.print("  [dim]All speakers in the dataset will be curated.[/dim]")
            raw = click.prompt("  Speaker IDs (comma-separated, Enter for all)", default="")
            result["speakers"] = raw.strip() or None

    # PHASE 4: SHOW CURATED SPEAKERS TO TRAIN
    elif 4 in phase_list:
        rows = _speakers_from_manifest(cfg)
        if rows:
            result["speakers"] = _speaker_picker(rows, "Curated Speakers — Phase 4 (Training)")
        else:
            console.print()
            console.print("  [dim]No curated speakers found. Run Phase 3 first.[/dim]")
            raw = click.prompt("  Speaker IDs (comma-separated, Enter for all)", default="")
            result["speakers"] = raw.strip() or None

    # PHASE 6: PICK WHICH SPOOFED SPEAKER FOLDERS TO ENHANCE
    if 6 in phase_list:
        spoofed_dir = Path(cfg["paths"]["output"]["spoofed"])
        if spoofed_dir.exists():
            speaker_dirs = sorted(d for d in spoofed_dir.iterdir() if d.is_dir())
            if speaker_dirs:
                rows = [
                    {"speaker_id": d.name, "wav_count": len(list(d.glob("*.wav")))}
                    for d in speaker_dirs
                ]
                result["speakers"] = _speaker_picker(rows, "Spoofed Speakers — Phase 6 (Enhancement)")

    # PHASE 5: PICK TTS SOURCE FILES AND SPEAKER MODELS INDEPENDENTLY
    if 5 in phase_list:
        tts_data = _tts_files_for_phase5(cfg)
        if tts_data:
            stems, src_dir = _tts_picker(tts_data)
            result["tts_stems"] = stems
            result["source_dir"] = src_dir
        else:
            console.print()
            console.print("  [dim]No TTS wav files found. Run Phase 2 first.[/dim]")

        model_rows = _models_for_phase5(cfg)
        if model_rows:
            result["speakers"] = _picker(
                model_rows, id_col="speaker_id",
                title="Speaker Models — Phase 5 (Inference)",
                prompt="Speaker selection",
            )
        else:
            console.print()
            console.print("  [dim]No trained models found. Run Phase 4 first.[/dim]")

    return result


@click.group(invoke_without_command=True)
@click.option("--config", "-c", default="config.yaml", show_default=True, help="Path to config.yaml")
@click.option("--verbose", "-v", is_flag=True, default=False, help="Enable debug logging")
@click.pass_context
def cli(ctx: click.Context, config: str, verbose: bool) -> None:
    """Isan-Spoofed: Thai-Isan voice conversion pipeline for ASV anti-spoofing research."""
    ctx.ensure_object(dict)
    cfg_path = Path(config)
    if not cfg_path.exists():
        raise click.ClickException(f"Config file not found: {cfg_path}")
    ctx.obj["config"] = _load_config(cfg_path)
    ctx.obj["verbose"] = verbose
    _configure_logging(Path(ctx.obj["config"]["paths"]["logs"]), verbose)

    # LAUNCH INTERACTIVE MENU WHEN NO SUBCOMMAND GIVEN
    if ctx.invoked_subcommand is None:
        _run_menu(ctx.obj["config"])


@cli.command()
@click.option("--mode", "-m", default=None,
              type=click.Choice(["tts", "rvc", "full"]),
              help="Mode shortcut: tts=phases 1-2, rvc=phases 3-4, full=phases 0-5. Overrides --phases.")
@click.option("--phases", "-p", default=None,
              help="Comma-separated phases to run, e.g. 1,3,5. Default: 1,2,3,4,5")
@click.option("--speakers", "-s", default=None,
              help="Comma-separated speaker IDs, e.g. f_091,m_023. Omit to process all.")
@click.option("--text", "-t", default=None,
              help="(Phase 1) Thai-Isan text to normalize, passed directly as a string.")
@click.option("--input-path", "-i", default=None,
              help="(Phase 1) Path to a .txt file or folder of .txt files to normalize.")
@click.option("--tts-files", default=None,
              help="(Phase 5) Comma-separated TTS file stems to convert, e.g. utterance_0001,utterance_0002.")
@click.option("--source-dir", default=None,
              help="(Phase 5) Folder of wav files to use as source instead of data/tts_output/.")
@click.pass_context
def run(
    ctx: click.Context,
    mode: Optional[str],
    phases: Optional[str],
    speakers: Optional[str],
    text: Optional[str],
    input_path: Optional[str],
    tts_files: Optional[str],
    source_dir: Optional[str],
) -> None:
    """Run the pipeline. Use --mode for shortcuts or --phases for fine control."""
    cfg = ctx.obj["config"]

    # RESOLVE PHASE LIST FROM MODE OR EXPLICIT --PHASES
    if mode is not None:
        phase_list = _MODE_PHASES[mode]
    elif phases is not None:
        phase_list = [int(p.strip()) for p in phases.split(",")]
    else:
        phase_list = [1, 2, 3, 4, 5]

    invalid = [p for p in phase_list if p not in _PHASES]
    if invalid:
        raise click.ClickException(f"Invalid phase(s): {invalid}. Valid: 0–5")

    # SHOW PICKERS IF NEEDED AND FLAGS NOT GIVEN
    if speakers is None:
        if 3 in phase_list:
            rows = _speakers_from_xlsx(cfg) or _speakers_from_raw_dataset(cfg)
            if rows:
                speakers = _speaker_picker(rows, "Available Speakers — Phase 3 (Curation)")
        elif 4 in phase_list:
            rows = _speakers_from_manifest(cfg)
            if rows:
                speakers = _speaker_picker(rows, "Curated Speakers — Phase 4 (Training)")

    if 5 in phase_list:
        if tts_files is None and source_dir is None:
            tts_data = _tts_files_for_phase5(cfg)
            if tts_data:
                tts_files, source_dir = _tts_picker(tts_data)
        if speakers is None:
            model_rows = _models_for_phase5(cfg)
            if model_rows:
                speakers = _picker(model_rows, id_col="speaker_id",
                                   title="Speaker Models — Phase 5 (Inference)",
                                   prompt="Speaker selection")

    extra = {"text": text, "input_path": input_path, "tts_stems": tts_files, "source_dir": source_dir}
    _execute_phases(cfg, phase_list, speakers, extra=extra)


@cli.command()
@click.pass_context
def augment(ctx: click.Context) -> None:
    """Apply data augmentation to the spoofed output corpus."""
    from isan_spoofed import augment as aug_module

    cfg = ctx.obj["config"]
    logger.info("Starting data augmentation")
    aug_module.run(cfg)
    logger.success("Augmentation complete")


@cli.command()
@click.pass_context
def status(ctx: click.Context) -> None:
    """Show current pipeline artifact counts."""
    cfg = ctx.obj["config"]
    counts = _artifact_counts(cfg)

    table = Table(title="Artifact Status", show_header=True, header_style="bold cyan")
    table.add_column("Artifact")
    table.add_column("Count", justify="right")
    for label, count in counts.items():
        table.add_row(label, str(count))

    console.print(table)


if __name__ == "__main__":
    cli()
